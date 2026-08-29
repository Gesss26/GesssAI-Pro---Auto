import json
import requests
from datetime import datetime
from typing import List, Dict
import os
import sys
import subprocess
import shutil
import time
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl non installato. Installa con: pip install openpyxl")

# ================================================================
# ===== CONFIGURAZIONE =====
# ================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Cartelle di destinazione
EXCEL_FOLDER = os.path.join(BASE_DIR, "excel")
JSON_FOLDER = os.path.join(BASE_DIR, "json")

# Crea le cartelle se non esistono
os.makedirs(EXCEL_FOLDER, exist_ok=True)
os.makedirs(JSON_FOLDER, exist_ok=True)

# ===== LISTA COMPLETA CAMPIONATI =====
LEAGUES = [
    {'name': 'Liga Profesional Argentina', 'url': 'https://www.matchesio.com/it/competition/liga-profesional-argentina-ar/export/json/'},
    {'name': 'Jupiler Pro League', 'url': 'https://www.matchesio.com/it/competition/jupiler-pro-league-be/export/json/'},
    {'name': 'Brasileirão Serie A', 'url': 'https://www.matchesio.com/it/competition/serie-a-br/export/json/'},
    {'name': 'Super League Grecia', 'url': 'https://www.matchesio.com/it/competition/super-league/export/json/'},
    {'name': 'Premier League', 'url': 'https://www.matchesio.com/it/competition/premier-league-gb-eng/export/json/'},
    {'name': 'EFL Championship', 'url': 'https://www.matchesio.com/it/competition/championship-gb-eng/export/json/'},
    {'name': 'Eredivisie', 'url': 'https://www.matchesio.com/it/competition/eredivisie-nl/export/json/'},
    {'name': 'Ligue 1', 'url': 'https://www.matchesio.com/it/competition/ligue-1-fr/export/json/'},
    {'name': 'Ligue 2', 'url': 'https://www.matchesio.com/it/competition/ligue-2-fr/export/json/'},
    {'name': 'Bundesliga', 'url': 'https://www.matchesio.com/it/competition/bundesliga-de/export/json/'},
    {'name': '2. Bundesliga', 'url': 'https://www.matchesio.com/it/competition/2-bundesliga-de/export/json/'},
    {'name': 'J1 League', 'url': 'https://www.matchesio.com/it/competition/j1-league/export/json/'},
    {'name': 'Serie A', 'url': 'https://www.matchesio.com/it/competition/serie-a-it/export/json/'},
    {'name': 'Serie B', 'url': 'https://www.matchesio.com/it/competition/serie-b-it/export/json/'},
    {'name': 'Serie C - Girone A', 'url': 'https://www.matchesio.com/it/competition/serie-c-girone-a-it/export/json/'},
    {'name': 'Serie C - Girone B', 'url': 'https://www.matchesio.com/it/competition/serie-c-girone-b-it/export/json/'},
    {'name': 'Serie C - Girone C', 'url': 'https://www.matchesio.com/it/competition/serie-c-girone-c-it/export/json/'},
    {'name': 'K League 1', 'url': 'https://www.matchesio.com/it/competition/k-league/export/json/'},
    {'name': 'Eerste Divisie', 'url': 'https://www.matchesio.com/it/competition/eerste-divisie-nl/export/json/'},
    {'name': 'Primeira Liga', 'url': 'https://www.matchesio.com/it/competition/primeira-liga-pt/export/json/'},
    {'name': 'Scottish Premiership', 'url': 'https://www.matchesio.com/it/competition/premiership-gb-sct/export/json/'},
    {'name': 'La Liga', 'url': 'https://www.matchesio.com/it/competition/la-liga-es/export/json/'},
    {'name': 'Segunda División', 'url': 'https://www.matchesio.com/it/competition/segunda-division-es/export/json/'},
    {'name': 'Süper Lig', 'url': 'https://www.matchesio.com/it/competition/super-lig-tr/export/json/'},
    {'name': 'Major League Soccer', 'url': 'https://www.matchesio.com/it/competition/major-league-soccer-us/export/json/'},
]

# ================================================================
# ===== CARICA LISTA NOMI SQUADRE DA FILE =====
# ================================================================

def load_team_names_from_file():
    """Carica i nomi delle squadre dal file lista nomi squadre.txt"""
    team_names = {}
    current_league = None
    
    try:
        file_path = os.path.join(BASE_DIR, 'lista nomi squadre.txt')
        
        if not os.path.exists(file_path):
            print(f"⚠️ File 'lista nomi squadre.txt' non trovato in: {file_path}")
            return {}
        
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        for line in lines:
            line = line.strip()
            
            # Salta linee vuote
            if not line:
                continue
            
            # Rileva il nome del campionato (es. "### Argentina – Liga Profesional")
            if line.startswith('###'):
                # Estrai il nome del campionato
                match = re.search(r'###\s*(.+?)\s*$', line)
                if match:
                    current_league = match.group(1).strip()
                    # Rimuovi eventuali "–" o "-"
                    current_league = re.sub(r'\s*[–\-]\s*', ' - ', current_league)
                    if current_league not in team_names:
                        team_names[current_league] = []
                continue
            
            # Rileva i nomi delle squadre (linee che iniziano con "* ")
            if line.startswith('*'):
                team_name = line[1:].strip()
                if current_league and team_name:
                    # Evita duplicati
                    if team_name not in team_names[current_league]:
                        team_names[current_league].append(team_name)
        
        # Stampa statistiche
        print(f"📋 Caricate {len(team_names)} campionati dal file 'lista nomi squadre.txt':")
        for league, teams in team_names.items():
            print(f"   • {league}: {len(teams)} squadre")
        
        return team_names
        
    except Exception as e:
        print(f"❌ Errore nel caricamento del file 'lista nomi squadre.txt': {e}")
        return {}

# ================================================================
# ===== CARICA LISTA SOSTITUZIONI DA FILE =====
# ================================================================

def load_team_name_mappings():
    """
    Carica le mappature dei nomi delle squadre dal file team_name_mappings.txt
    Formato: nome_originale|nome_corretto
    """
    mappings = {}
    
    try:
        file_path = os.path.join(BASE_DIR, 'team_name_mappings.txt')
        
        if not os.path.exists(file_path):
            return {}
        
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if '|' in line:
                    parts = line.split('|')
                    if len(parts) == 2:
                        mappings[parts[0].strip()] = parts[1].strip()
        
        return mappings
        
    except Exception as e:
        print(f"⚠️ Errore nel caricamento delle mappature: {e}")
        return {}

# ================================================================
# ===== FUNZIONI DI CORREZIONE NOMI =====
# ================================================================

def find_team_in_list(team_name: str, team_list: List[str]) -> str:
    """
    Cerca un nome di squadra in una lista, restituendo il nome corretto.
    Usa il matching fuzzy per trovare corrispondenze.
    """
    if not team_name or not team_list:
        return team_name
    
    team_lower = team_name.lower().strip()
    
    # Prima cerca una corrispondenza esatta (case insensitive)
    for correct_name in team_list:
        if correct_name.lower() == team_lower:
            return correct_name
    
    # Poi cerca corrispondenze parziali
    # Rimuovi prefissi comuni come "FC ", "SC ", "SS "
    prefixes = ['fc ', 'sc ', 'ss ', 'as ', 'ac ', 'us ', 'cd ', 'cf ', 'de ']
    clean_team = team_lower
    for prefix in prefixes:
        if clean_team.startswith(prefix):
            clean_team = clean_team[len(prefix):]
            break
    
    # Rimuovi suffissi comuni come " fc", " sc", " ss"
    suffixes = [' fc', ' sc', ' ss', ' as', ' ac', ' cf']
    for suffix in suffixes:
        if clean_team.endswith(suffix):
            clean_team = clean_team[:-len(suffix)]
            break
    
    clean_team = clean_team.strip()
    
    # Cerca corrispondenze parziali
    best_match = None
    best_score = 0
    
    for correct_name in team_list:
        correct_lower = correct_name.lower()
        
        # Corrispondenza esatta dopo la pulizia
        if clean_team == correct_lower.replace('fc ', '').replace('sc ', '').strip():
            return correct_name
        
        # Corrispondenza parziale (contiene)
        if clean_team in correct_lower or correct_lower in clean_team:
            score = len(clean_team) / max(len(correct_lower), 1)
            if score > best_score:
                best_score = score
                best_match = correct_name
    
    # Se la corrispondenza è sufficientemente buona (oltre il 60%)
    if best_match and best_score > 0.6:
        return best_match
    
    return team_name

def correct_team_names_for_league(matches: List[Dict], league_name: str, team_names_dict: Dict) -> List[Dict]:
    """
    Corregge i nomi delle squadre per un campionato usando la lista fornita.
    """
    if not matches:
        return matches
    
    # Trova il campionato nella lista
    league_key = None
    for key in team_names_dict.keys():
        if key.lower() == league_name.lower():
            league_key = key
            break
        # Controlla se il nome del campionato è contenuto
        if league_name.lower() in key.lower() or key.lower() in league_name.lower():
            league_key = key
            break
    
    if not league_key:
        # Se non trovato, prova a cercare per similarità
        for key in team_names_dict.keys():
            # Rimuovi parole comuni per il confronto
            key_clean = re.sub(r'(campionato|league|serie|division|liga|super|premier|championship|professional|pro|league|1|2|3)', '', key.lower())
            league_clean = re.sub(r'(campionato|league|serie|division|liga|super|premier|championship|professional|pro|league|1|2|3)', '', league_name.lower())
            if key_clean.strip() and league_clean.strip() and (key_clean in league_clean or league_clean in key_clean):
                league_key = key
                break
    
    if not league_key:
        print(f"   ⚠️ Campionato '{league_name}' non trovato nella lista nomi. Uso nomi originali.")
        return matches
    
    team_list = team_names_dict[league_key]
    corrected_count = 0
    
    for match in matches:
        # Corregge squadra casa
        if match.get('squadra_casa'):
            original = match['squadra_casa']
            corrected = find_team_in_list(original, team_list)
            if corrected != original:
                match['squadra_casa'] = corrected
                corrected_count += 1
        
        # Corregge squadra ospite
        if match.get('squadra_ospite'):
            original = match['squadra_ospite']
            corrected = find_team_in_list(original, team_list)
            if corrected != original:
                match['squadra_ospite'] = corrected
                corrected_count += 1
    
    if corrected_count > 0:
        print(f"   ✅ Corretti {corrected_count} nomi squadre per {league_key}")
    
    return matches

# ================================================================
# ===== FUNZIONI DI UTILITÀ =====
# ================================================================

def convert_date_to_italian(date_str: str) -> str:
    """Converte una data da YYYY-MM-DD a DD/MM/YYYY."""
    if not date_str:
        return ''
    
    date_str = str(date_str)
    
    if re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
        return date_str
    
    match = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
    if match:
        return f"{match.group(3)}/{match.group(2)}/{match.group(1)}"
    
    return date_str

def parse_date_for_sorting(date_str: str) -> str:
    """Converte DD/MM/YYYY in YYYY-MM-DD per l'ordinamento."""
    if not date_str:
        return '9999-99-99'
    
    match = re.search(r'(\d{2})/(\d{2})/(\d{4})', date_str)
    if match:
        return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
    
    return date_str

def parse_matches_from_json(data, league_name: str) -> List[Dict]:
    """Estrae le partite dal JSON."""
    matches = []
    
    if not isinstance(data, list):
        print(f"   ⚠️ Il JSON non è una lista, è {type(data)}")
        return []
    
    print(f"   📊 Trovate {len(data)} partite")
    
    for match in data:
        if not isinstance(match, dict):
            continue
        
        # === SQUADRE - NOMI ORIGINALI DA MATCHESIO ===
        home_team = match.get('homeTeam', '') or match.get('home_team', '') or ''
        away_team = match.get('awayTeam', '') or match.get('away_team', '') or ''
        
        date_raw = match.get('date', '')
        date_str = convert_date_to_italian(date_raw)
        
        time_str = match.get('time', '')
        matchday = match.get('matchday', '')
        
        result_str = match.get('result', '')
        home_score = ''
        away_score = ''
        result = ''
        
        if result_str:
            result_clean = result_str.replace('–', '-')
            if '-' in result_clean:
                parts = result_clean.split('-')
                if len(parts) == 2:
                    home_score = parts[0].strip()
                    away_score = parts[1].strip()
                    result = f"{home_score}–{away_score}"
            else:
                result = result_str
        
        status = match.get('status', '').lower()
        
        if status in ['giocata', 'played', 'finished', 'complete', 'completed', 'ft']:
            status_ita = 'Giocata'
        elif status in ['da giocare', 'scheduled', 'upcoming', 'future', 'not started', '']:
            if home_score and away_score:
                status_ita = 'Giocata'
            else:
                status_ita = 'Futura'
        elif status in ['in corso', 'live', 'in progress']:
            status_ita = 'In corso'
        elif status in ['posticipata', 'postponed']:
            status_ita = 'Posticipata'
        else:
            status_ita = 'Futura'
        
        city = match.get('city', '')
        stadium = match.get('stadium', '')
        sort_date = parse_date_for_sorting(date_str)
        
        matches.append({
            'campionato': league_name,
            'data': date_str,
            'ora': time_str,
            'giornata': str(matchday),
            'squadra_casa': home_team,
            'squadra_ospite': away_team,
            'risultato': result,
            'gol_casa': home_score,
            'gol_ospite': away_score,
            'citta': city,
            'stadio': stadium,
            'stato': status_ita,
            '_sort_date': sort_date
        })
    
    return matches

def sort_and_clean_matches(all_matches: List[Dict]) -> List[Dict]:
    """
    Ordina le partite per data e RIMUOVE il campo _sort_date.
    """
    # Rimuovi duplicati
    unique_matches = []
    seen = set()
    for match in all_matches:
        key = (match['campionato'], match['data'], match['ora'], 
               match['squadra_casa'], match['squadra_ospite'])
        if key not in seen:
            seen.add(key)
            unique_matches.append(match)
    
    # Ordina per data
    unique_matches.sort(key=lambda x: (x.get('_sort_date', '9999-99-99'), x['campionato']))
    
    # RIMUOVI IL CAMPO _sort_date DA TUTTE LE PARTITE
    for match in unique_matches:
        if '_sort_date' in match:
            del match['_sort_date']
    
    return unique_matches

def fetch_league_json(url: str, league_name: str) -> List[Dict]:
    """Scarica il JSON da un campionato."""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        response = requests.get(url, timeout=30, headers=headers)
        response.raise_for_status()
        
        data = response.json()
        matches = parse_matches_from_json(data, league_name)
        return matches
        
    except Exception as e:
        print(f"❌ Errore nel download di {league_name}: {e}")
        return []

# ================================================================
# ===== SALVATAGGIO FILE =====
# ================================================================

def save_excel(all_matches: List[Dict]) -> str:
    """Salva i dati in Excel (.xlsx) nella cartella excel/."""
    if not all_matches or not EXCEL_AVAILABLE:
        return None
    
    try:
        # Ordina e pulisci i dati
        all_matches = sort_and_clean_matches(all_matches)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Calcio"
        
        headers = ['Campionato', 'Data', 'Ora', 'Giornata', 'Squadra Casa', 'Squadra Ospite', 
                   'Risultato', 'Gol Casa', 'Gol Ospite', 'Città', 'Stadio', 'Stato']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row, match in enumerate(all_matches, 2):
            ws.cell(row=row, column=1, value=match.get('campionato', ''))
            ws.cell(row=row, column=2, value=match.get('data', ''))
            ws.cell(row=row, column=3, value=match.get('ora', ''))
            ws.cell(row=row, column=4, value=match.get('giornata', ''))
            ws.cell(row=row, column=5, value=match.get('squadra_casa', ''))
            ws.cell(row=row, column=6, value=match.get('squadra_ospite', ''))
            ws.cell(row=row, column=7, value=match.get('risultato', ''))
            ws.cell(row=row, column=8, value=match.get('gol_casa', ''))
            ws.cell(row=row, column=9, value=match.get('gol_ospite', ''))
            ws.cell(row=row, column=10, value=match.get('citta', ''))
            ws.cell(row=row, column=11, value=match.get('stadio', ''))
            ws.cell(row=row, column=12, value=match.get('stato', ''))
        
        column_widths = [25, 15, 10, 10, 22, 22, 12, 10, 10, 18, 25, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        excel_path = os.path.join(EXCEL_FOLDER, 'GesssAI_Input.xlsx')
        wb.save(excel_path)
        wb.close()
        
        print(f"   ✅ GesssAI_Input.xlsx salvato in: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"   ❌ Errore nel salvataggio Excel: {e}")
        return None

def save_json(all_matches: List[Dict]):
    """Salva i dati in JSON:
       - json/GesssAI_Input.json
       - matches.json (nella root)
    """
    if not all_matches:
        return None, None
    
    try:
        # Ordina e pulisci i dati (rimuove _sort_date)
        all_matches = sort_and_clean_matches(all_matches)
        
        # ===== 1. SALVA IN json/GesssAI_Input.json =====
        json_path = os.path.join(JSON_FOLDER, 'GesssAI_Input.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(all_matches, f, ensure_ascii=False, indent=2)
        print(f"   ✅ GesssAI_Input.json salvato in: {json_path}")
        
        # ===== 2. SALVA IN ROOT matches.json =====
        root_json_path = os.path.join(BASE_DIR, 'matches.json')
        with open(root_json_path, 'w', encoding='utf-8') as f:
            json.dump(all_matches, f, ensure_ascii=False, indent=2)
        print(f"   ✅ matches.json salvato in: {root_json_path}")
        
        return json_path, root_json_path
        
    except Exception as e:
        print(f"   ❌ Errore nel salvataggio JSON: {e}")
        return None, None

# ================================================================
# ===== FUNZIONE PER INVIARE SU GITHUB =====
# ================================================================

def push_to_github():
    """Esegue git add, commit e push dei file su GitHub."""
    try:
        current_dir = os.getcwd()
        os.chdir(BASE_DIR)
        
        print("\n🔄 Eseguo git add...")
        
        # Aggiungi i file uno per uno
        result = subprocess.run(['git', 'add', 'excel/GesssAI_Input.xlsx'], capture_output=True, text=True)
        if result.returncode != 0:
            print(f"   ⚠️ Errore git add excel: {result.stderr}")
        
        result = subprocess.run(['git', 'add', 'json/GesssAI_Input.json'], capture_output=True, text=True)
        if result.returncode != 0:
            print(f"   ⚠️ Errore git add json: {result.stderr}")
        
        result = subprocess.run(['git', 'add', 'matches.json'], capture_output=True, text=True)
        if result.returncode != 0:
            print(f"   ⚠️ Errore git add matches: {result.stderr}")
        
        print("   ✅ File aggiunti")
        
        print("🔄 Eseguo git commit...")
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
        commit_msg = f"Aggiornati dati calcio da matchesio.com ({timestamp})"
        commit_result = subprocess.run(['git', 'commit', '-m', commit_msg], 
                                     capture_output=True, text=True)
        
        if "nothing to commit" in commit_result.stdout:
            print("   ℹ️ Nessuna modifica da committare")
            os.chdir(current_dir)
            return True
        
        print("   ✅ Commit effettuato")
        
        print("🔄 Eseguo git push...")
        result = subprocess.run(['git', 'push', 'origin', 'master'], 
                              capture_output=True, text=True)
        
        if result.returncode == 0:
            print("\n✅ FILE INVIATI CON SUCCESSO SU GITHUB!")
            print(f"   📁 https://github.com/Gesss26/GesssAI-Pro---Auto")
            os.chdir(current_dir)
            return True
        else:
            print(f"\n❌ Errore durante il push: {result.stderr}")
            os.chdir(current_dir)
            return False
        
    except Exception as e:
        print(f"\n❌ Errore: {e}")
        return False

# ================================================================
# ===== MAIN =====
# ================================================================

def main():
    print("="*70)
    print("🚀 Avvio download calendari calcio da matchesio.com...")
    print(f"📋 {len(LEAGUES)} campionati da scaricare")
    print("="*70)
    print()
    
    # ===== CARICA I NOMI DELLE SQUADRE =====
    print("📂 Caricamento lista nomi squadre da 'lista nomi squadre.txt'...")
    team_names_dict = load_team_names_from_file()
    print()
    
    all_matches = []
    errors = []
    
    for i, league in enumerate(LEAGUES, 1):
        league_name = league['name']
        print(f"⏳ [{i:2d}/{len(LEAGUES)}] Scaricando {league_name}...")
        
        matches = fetch_league_json(league['url'], league_name)
        
        if matches:
            # ===== CORREGGI I NOMI DELLE SQUADRE =====
            matches = correct_team_names_for_league(matches, league_name, team_names_dict)
            
            all_matches.extend(matches)
            giocate = sum(1 for m in matches if m['stato'] == 'Giocata')
            future = sum(1 for m in matches if m['stato'] == 'Futura')
            print(f"   ✅ {len(matches)} partite ({giocate} giocate, {future} future)")
        else:
            errors.append(league_name)
            print(f"   ❌ Nessuna partita trovata")
        
        time.sleep(0.5)
    
    if all_matches:
        print("\n💾 Salvataggio file...")
        
        excel_path = save_excel(all_matches)
        json_path, root_json_path = save_json(all_matches)
        
        if excel_path:
            print(f"\n📁 Cartella excel: {EXCEL_FOLDER}")
        if json_path:
            print(f"📁 Cartella json: {JSON_FOLDER}")
        if root_json_path:
            print(f"📁 Root: matches.json")
        
        # Statistiche
        giocate = sum(1 for m in all_matches if m['stato'] == 'Giocata')
        future = sum(1 for m in all_matches if m['stato'] == 'Futura')
        
        print("\n" + "="*70)
        print("📊 STATISTICHE FINALI")
        print("="*70)
        print(f"   • Partite totali: {len(all_matches):,}")
        print(f"   • Giocate: {giocate:,}")
        print(f"   • Future: {future:,}")
        print("="*70)
        
        print("\n📤 INVIO SU GITHUB...")
        push_to_github()
        
    else:
        print("\n❌ Nessuna partita scaricata.")
    
    print("\n" + "="*70)
    print("🏁 Script terminato")
    print("="*70)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ ERRORE: {e}")
        import traceback
        traceback.print_exc()
    
    # 🔥 TENIAMO APERTA LA FINESTRA 🔥
    input("\n🔄 Premi ENTER per uscire...")